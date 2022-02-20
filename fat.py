from openpyxl import load_workbook

vendedores = [("Felipe Santana","felipe.santana@msfoods.agr.br"),("Marcos Manso","marcos.manso@msfoods.agr.br"),("Andrei Almeida","andrei.almeida@msfoods.agr.br"),
              ("Anderson Brito","anderson.brito@msfoods.agr.br"),("Felipe Fagundes","felipe.correa@msfoods.agr.br"),("Leandro Tenalia","leandro.tenalia@msfoods.agr.br")]

wb = load_workbook(filename="C:/Users/Tecnofoods/Google Drive/Faturamento/Faturamento 2019 2020.xlsx", data_only=True)
sheet = wb["Base de Dados"]

j = 0

for i in vendedores:

    wb2 = load_workbook(filename=f"C:/Users/Tecnofoods/Desktop/Resumos/ResumoCoord/{vendedores[j][0]}.xlsx")
    sheet3 = wb2["Geral"]

    lin = 3

    for row in sheet.values:
    
        if row[0] == vendedores[j][0] and row[4] == 2021:
            sheet3[f"A{lin}"].value = row[15]
            sheet3[f"B{lin}"].value = row[1]
            sheet3[f"C{lin}"].value = row[2]
            sheet3[f"D{lin}"].value = row[5]
            sheet3[f"E{lin}"].value = row[7]
            sheet3[f"F{lin}"].value = row[8]
            sheet3[f"G{lin}"].value = row[9]
            sheet3[f"H{lin}"].value = row[10]
            sheet3[f"I{lin}"].value = row[11]
            sheet3[f"J{lin}"].value = row[12]
            sheet3[f"K{lin}"].value = row[13]
            sheet3[f"L{lin}"].value = row[28]
            lin = lin + 1
    
    

    wb2.save(filename=f"C:/Users/Tecnofoods/Desktop/Resumos/ResumoCoord/{vendedores[j][0]}.xlsx")

    j = j + 1

lin = 3

wb2 = load_workbook(filename=f"C:/Users/Tecnofoods/Desktop/Resumos/ResumoCoord/Resumo Geral.xlsx")
sheet3 = wb2["Geral"]

for row in sheet.values:
    
    if row[4] == 2021:
        sheet3[f"A{lin}"].value = row[0]
        sheet3[f"B{lin}"].value = row[15]
        sheet3[f"C{lin}"].value = row[1]
        sheet3[f"D{lin}"].value = row[2]
        sheet3[f"E{lin}"].value = row[5]
        sheet3[f"F{lin}"].value = row[7]
        sheet3[f"G{lin}"].value = row[8]
        sheet3[f"H{lin}"].value = row[9]
        sheet3[f"I{lin}"].value = row[10]
        sheet3[f"J{lin}"].value = row[11]
        sheet3[f"K{lin}"].value = row[12]
        sheet3[f"L{lin}"].value = row[13]
        lin = lin + 1
    
    

wb2.save(filename=f"C:/Users/Tecnofoods/Desktop/Resumos/ResumoCoord/Resumo Geral.xlsx")






    