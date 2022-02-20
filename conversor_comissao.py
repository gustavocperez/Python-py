import tabula
from openpyxl import load_workbook

df = tabula.read_pdf('C:/Users/TEMP/Documents/TECNOFOODS/LMT/COMISSAO-LMT REPRESENTACOES LTDA - ME.pdf')

nome_arquivo = 'comissao.xlsx'

for i in df:
    i.to_excel(nome_arquivo)

workbook = load_workbook(filename=nome_arquivo)
ws = workbook.active

ws.delete_cols(1)
ws.delete_cols(5)
ws.delete_cols(7)

ws.delete_rows(1)

j = 1
i = 1

for cell in ws["D"]:
    dado = ws[f"D{j}"].value
    try: 
        if (dado.endswith("Total")) == True:
            ws.delete_rows(j)
    except:
        print("")
    finally:
        j = j + 1

j = 1

for cell in ws["A"]:
    dado = ws[f"A{j}"].value
    try:
        if dado.endswith("Total"):
            ws.delete_rows(j)
    except: print("")
    finally: j = j + 1

j = 1

for cell in ws["B"]:
    dado = ws[f"B{j}"].value
    if dado is None:
        ws[f"B{j}"].value = ws[f"B{j-1}"].value
        ws[f"A{j}"].value = ws[f"A{j-1}"].value
    j = j + 1

j = 1

for cell in ws["C"]:
    dado = ws[f"C{j}"].value
    if dado == None:
        ws[f"C{j}"].value = ws[f"C{j-1}"].value
    j = j + 1

j = 1

for cell in ws["D"]:
    dado = ws[f"D{j}"].value
    if dado == None:
        ws[f"D{j}"].value = ws[f"D{j-1}"].value
    j = j + 1

for cell in ws["A"]:
    dado = ws[f"A{j}"].value
    try: 
        if (dado.endswitch("Geral")) == True:
            ws.delete_rows(j)
    except: print("")
    finally: j = j + 1

j = 1

for cell in ws["A"]:
    dado = ws[f"E{j}"].value
    if dado == None:
        ws.delete_rows(j)
    j = j + 1


workbook.save(filename=nome_arquivo)